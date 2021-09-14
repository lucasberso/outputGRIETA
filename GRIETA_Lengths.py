#
# Authors:
# Lucas Bernacer Soriano,
# Javier Vela
#
# Copyright (c) 2021, Capgemini Engineering
#


import os
from typing import List, Any
import numpy as np
from path import Path
import pandas as pd
from matplotlib import pyplot as plt
from openpyxl import load_workbook
from wrappers import read_table_horiz
from prettytable import PrettyTable

class GRIETA_Critical_Lengths():
    """
    Clase creada para leer el input del archivo Excel
    """
    def __init__(self, Excel_file, folder_data):
        # INICIALIZACIÓN DE LA CARPETA
        # Excel_file: Archivo plantilla de Excel
        # folder_data: Carpeta donde se encuentran los diferentes archivos .num

        # Definición de archivos propios de la carpeta
        self.folder_data = folder_data
        self.Excel_file = Excel_file

        # Abrir libro Excel
        self.book = load_workbook(self.Excel_file, data_only = True)

        # Guardar datos comunes del problema como archivos propios de la clase
        self.files_missions = []
        for row in range(7,12):
            self.files_missions.append(self.book["INPUTS"].cell(row,2).value)
        self.Crack_Length_Calc_Method = self.book["INPUTS"].cell(14,2).value
        self.hole_diameter = self.book["INPUTS"].cell(17,2).value
        self.limit_stress = self.book["INPUTS"].cell(21,2).value
        self.load_case = self.book["INPUTS"].cell(22,2).value


    def Read_files(self):
        # FUNCION PARA LEER LOS DIFERENTES ARCHIVOS .num

        # Lectura de los archivos que hay en la carpeta
        list_files = os.listdir(self.folder_data)
        SR_files = []
        MR_files = []
        LR_files = []
        ULR_files = []
        MIX_files = []
        for file in list_files:
            for mission_file in self.files_missions:
                if mission_file != None:
                    # Se guardan los archivos .num de la misión especificada en la hoja Excel de input
                    if file[:len(mission_file)] == mission_file and file[-5:] == "num.1":
                        if mission_file[-2:] == "sr":
                            SR_files.append(file)
                        elif mission_file[-2:] == "mr":
                            MR_files.append(file)
                        elif mission_file[-2:] == "lr":
                            LR_files.append(file)
                        elif mission_file[-3:] == "ulr":
                            ULR_files.append(file)
                        elif mission_file[-3:] == "mix":
                            MIX_files.append(file)
        # Se archivan todos los archivos en una lista común
        All_missions = []
        All_missions.append(SR_files)
        All_missions.append(MR_files)
        All_missions.append(LR_files)
        All_missions.append(ULR_files)
        All_missions.append(MIX_files)

        # Leer los diferentes archivos para sacar todos los datos
        # Se hace un bucle a través de todos los archivos en All_missions
        content_mat = []
        for i in range(0, len(All_missions)):
            content_mat.append([])
            input_steps = 0
            for j in range(0, len(All_missions[i])):
                filename = self.folder_data + "/" + All_missions[i][j]
                with open(filename) as f:
                    content = f.readlines()
                    for line in content:
                        new_list_2 = []
                        new_list = line.split()
                        for item in new_list:
                            new_list_2.append(float(item))

                        # Una línea de ceros significa que hay un cambio de step
                        All_zeros = "Yes"
                        for t in range(0, len(new_list_2)):
                            if new_list_2[t] != 0:
                                All_zeros = "No"

                        if All_zeros == "No":
                            new_list_2.append(input_steps)
                            content_mat[i].append(new_list_2)
                        else:
                            input_steps = input_steps + 1
        content = content_mat
        # Reordenar los datos extraidos para que se puedan poner en formato dataframe de Pandas
        content_order = []
        for k in range(0, len(content)):
            content_order.append([])
            if len(content[k]) > 0:
                for j in range(0, len(content[k][0])):
                    content_order[k].append([])
                    for i in range(0, len(content[k])):
                        content_order[k][j].append(content[k][i][j])

        # Nombre de las columnas de los datos
        name_list = ["Crack a", "Crack b", "Crack c", "Crack d", "Cycles", "Flights", "Beta a", "Beta b", "Beta c",
                     "Beta d", "Klim ab", "Klim cd", "Input Steps"]

        # Se guardan todos los datos extraídos en formato dataframe de Pandas con la cabecera de name_list
        df_all = []
        for k in range(0, len(content_order)):
            df_all.append([])
            df_all[k] = pd.DataFrame()
            if len(content_order[k]) > 0:
                for i in range(0, len(name_list)):
                    df_all[k][name_list[i]] = content_order[k][i]

        return(df_all)


    def Obtain_Crack_Length(self, dataframe_all):
        # FUNCION PARA CALCULAR LA CRACK LENGTH DE TODOS LOS DATOS QUE HAN SIDO EXTRAIDOS

        # Bucle a través del dataframe con los datos
        for k in range(0, len(dataframe_all)):
            atotal_list = []
            for i in range(0, dataframe_all[k].shape[0]):
                # Si la grieta no es pasante todavía se suma el valor del diámetro del agujero
                if dataframe_all[k]["Crack b"][i] != 0 or dataframe_all[k]["Crack d"][i] != 0:
                    if self.Crack_Length_Calc_Method == "A":
                        atotal = dataframe_all[k]["Crack a"][i] + self.hole_diameter
                        atotal_list.append(atotal)
                    elif self.Crack_Length_Calc_Method == "C":
                        atotal = dataframe_all[k]["Crack c"][i] + self.hole_diameter
                        atotal_list.append(atotal)
                    elif self.Crack_Length_Calc_Method == "A+C":
                        atotal = dataframe_all[k]["Crack a"][i] + dataframe_all[k]["Crack c"][i] + self.hole_diameter
                        atotal_list.append(atotal)
                # Una vez la grieta es pasante no se suma el diámetro, ya lo considera el programa
                else:
                    if self.Crack_Length_Calc_Method == "A":
                        atotal = dataframe_all[k]["Crack a"][i]
                        atotal_list.append(atotal)
                    elif self.Crack_Length_Calc_Method == "C":
                        atotal = dataframe_all[k]["Crack c"][i]
                        atotal_list.append(atotal)
                    elif self.Crack_Length_Calc_Method == "A+C":
                        atotal = dataframe_all[k]["Crack a"][i] + dataframe_all[k]["Crack c"][i]
                        atotal_list.append(atotal)
            # Actualizar el dataframe para incluir la longitud de la grieta
            dataframe_all[k]['a total'] = atotal_list
        return(dataframe_all)


    def Plot_Crack_Length(self, dataframe_all):
        # FUNCION PARA PLOTEAR LAS LONGITUDES DE GRIETA VS CICLOS
        # No se usa en la herramienta final
        missions_list = ["SR", "MR", "LR", "ULR", "MIX"]
        legend_mat = []
        for k in range(0, len(dataframe_all)):
            if len(dataframe_all[k]) > 0:
                plt.plot(dataframe_all[k]["Flights"], dataframe_all[k]['a total'])
                legend_mat.append(missions_list[k])
        plt.legend(legend_mat)
        plt.xlabel("Number of Flights")
        plt.ylabel("Crack length (mm)")
        plt.show()


    def Compute_Critical_Crack_Lengths(self, df_all, output_folder, txt_name):
        # FUNCION PARA CALCULAR LAS LONGITUDES CRÍTICAS DE GRIETA MEDIANTE LOS DIFERENTES MÉTODOS

        # Lectura del Excel para saber qué criterios hay que aplicar
        Fracture_Mechs_criterion = self.book["INPUTS"].cell(24,2).value
        Net_sect_yield_criterion = self.book["INPUTS"].cell(25, 2).value
        Fast_crack_growth_criterion = self.book["INPUTS"].cell(26, 2).value
        Crit_crack_length_cons = self.book["INPUTS"].cell(27, 2).value

        row_FM = 25
        row_NS = 26
        row_FC = 27

        # Limpiar las celdas de resultados del Excel en caso de que haya algo escrito
        for row in range(row_FM, row_FC + 1):
            for col in range(8, 14):
                self.book["INPUTS"].cell(row, col).value = ""
        self.book["INPUTS"].cell(row_FM, 14).value = ""

        # Creación de listas para sacar cuáles son los máximos
        Global_max_lengths = []
        Global_min_lengths = []
        All_Crit_Lengths = []

        # Se va viendo para cada criterio si hay que hacer algún cálculo y se aplica el correspondiente
        if Fracture_Mechs_criterion == "KR curve":
            # Leer curva R del Excel
            row = 35
            col = 1
            KR_curve_list = []
            cont = 0
            while self.book["INPUTS"].cell(row, col).value != None:
                KR_curve_list.append([])
                KR_curve_list[cont].append(self.book["INPUTS"].cell(row, col).value)
                KR_curve_list[cont].append(self.book["INPUTS"].cell(row, col + 1).value)
                row = row + 1
                cont = cont + 1
            KR_curve = np.array(KR_curve_list)

            # Aplicar función que calcula la longitud de grieta a partir de la curva R
            Crit_Lengths_FM = self.KR_curve_calc(df_all, KR_curve)
            All_Crit_Lengths.append(Crit_Lengths_FM)

            # Cálculo de la máxima y mínima longitud de grieta y escribir en el Excel los resultados
            (Max_length_FM, Min_length_FM) = self.Write_Critical_Lengths(row_FM, Crit_Lengths_FM, Crit_crack_length_cons)

            Global_min_lengths.append(Min_length_FM)
            Global_max_lengths.append(Min_length_FM)

        elif Fracture_Mechs_criterion == "Residual strength":
            # Aplicar función que calcula la longitud de grieta a partir de resistencia residual
            Crit_Lengths_FM = self.Residual_Strength_calc(df_all)
            All_Crit_Lengths.append(Crit_Lengths_FM)

            # Cálculo de la máxima y mínima longitud de grieta y escribir en el Excel los resultados
            (Max_length_FM, Min_length_FM) = self.Write_Critical_Lengths(row_FM, Crit_Lengths_FM,
                                                                         Crit_crack_length_cons)
            Global_min_lengths.append(Min_length_FM)
            Global_max_lengths.append(Min_length_FM)

        else:
            # Si no hay ningún método seleccionado se dejan vacíos los huecos respecto a este criterio
            All_Crit_Lengths.append([])
            for i in range(0, 5):
                All_Crit_Lengths[0].append("")
            Global_min_lengths.append("")
            Global_max_lengths.append("")

        if Net_sect_yield_criterion == "Yes":
            A_total = self.book["INPUTS"].cell(29, 2).value
            Fty = self.book["INPUTS"].cell(30, 2).value

            # Aplicar función que calcula la longitud de grieta a partir de sección  neta
            Crit_Lengths_NSY = self.Net_Sec_Yield_calc(df_all, A_total, Fty)
            All_Crit_Lengths.append(Crit_Lengths_NSY)

            # Cálculo de la máxima y mínima longitud de grieta y escribir en el Excel los resultados
            (Max_length_NSY, Min_length_NSY) = self.Write_Critical_Lengths(row_NS, Crit_Lengths_NSY, Crit_crack_length_cons)

            Global_min_lengths.append(Min_length_NSY)
            Global_max_lengths.append(Min_length_NSY)

        else:
            # Si no hay ningún método seleccionado se dejan vacíos los huecos respecto a este criterio
            All_Crit_Lengths.append([])
            for i in range(0, 5):
                All_Crit_Lengths[1].append("")
            Global_min_lengths.append("")
            Global_max_lengths.append("")

        if Fast_crack_growth_criterion != "No":
            # Aplicar función que calcula la longitud de grieta a partir de crecimiento rápido grieta
            Crit_Lengths_CWC = self.Fast_Growth_Crack_calc(Fast_crack_growth_criterion, df_all)
            All_Crit_Lengths.append(Crit_Lengths_CWC)

            # Cálculo de la máxima y mínima longitud de grieta y escribir en el Excel los resultados
            (Max_length_CWC, Min_length_CWC) = self.Write_Critical_Lengths(row_FC, Crit_Lengths_CWC, Crit_crack_length_cons)

            Global_min_lengths.append(Min_length_CWC)
            Global_max_lengths.append(Min_length_CWC)
        else:
            # Si no hay ningún método seleccionado se dejan vacíos los huecos respecto a este criterio
            All_Crit_Lengths.append([])
            for i in range(0, 5):
                All_Crit_Lengths[2].append("")
            Global_min_lengths.append("")
            Global_max_lengths.append("")

        # Se hacen estos bucles para obtener cuales son las máximas/mínimas longitudes de grieta
        for i in range(0, len(Global_max_lengths)):
            if Global_max_lengths[i] != "":
                Max_length = Global_max_lengths[i]
                Min_length = Global_min_lengths[i]
                break
        for i in range(1, len(Global_max_lengths)):
            if Global_max_lengths[i] != "":
                if Global_max_lengths[i] > Max_length:
                    Max_length = Global_max_lengths[i]
            if Global_min_lengths[i] != "":
                if Global_min_lengths[i] < Min_length:
                    Min_length = Global_min_lengths[i]

        # Se escriben en el Excel los valores de máxima/mínima longitud de grieta
        # Dependiendo de si la aproximación es real o conservativa
        if Crit_crack_length_cons == "Real":
            self.book["INPUTS"].cell(25, 14).value = Max_length
        elif Crit_crack_length_cons == "Conservative":
            self.book["INPUTS"].cell(25, 14).value = Min_length

        # Se guarda el archivo Excel y se cierra
        self.book.save(self.Excel_file)

        # Función para escribir el output en un archivo txt
        self.Write_Critical_Lengths_txt(output_folder, txt_name, All_Crit_Lengths, Global_max_lengths,
                                        Global_min_lengths, Crit_crack_length_cons)


    def KR_curve_calc(self, df_all, KR_curve):
        # FUNCIÓN PARA CALCULAR LA LONGITUD CRÍTICA CON LA CURVA R

        # Mínimo incremento para buscar en la curva R
        # Este valor es debido a que a veces en la curva R si hay una parte recta se da directamente el valor inicial
        # y el final, este incremento hace que se busque cada x ciclos si hay tangencia en esos puntos
        des_inc = 5
        KR_curve_comp = []
        KR_curve_comp.append([])
        KR_curve_comp.append([])
        KR_curve_comp[0].append(KR_curve[0,0])
        KR_curve_comp[1].append(KR_curve[0,1])
        # Obtención de los diferentes puntos en la curva R cuando el incremento sea muy grande
        for i in range(1, len(KR_curve)):
            if KR_curve[i, 0] - KR_curve[i - 1, 0] <= des_inc:
                KR_curve_comp[0].append(KR_curve[i, 0])
                KR_curve_comp[1].append(KR_curve[i, 1])
            else:
                increments = round((KR_curve[i, 0] - KR_curve[i - 1, 0]) / des_inc)
                for j in range(0, increments):
                    a_int = KR_curve[i - 1, 0] + des_inc * (j + 1)
                    K_int = KR_curve[i - 1, 1] + (KR_curve[i, 1] - KR_curve[i - 1, 1])/(KR_curve[i, 0] - KR_curve[i - 1, 0]) * (a_int - KR_curve[i - 1, 0])
                    KR_curve_comp[0].append(a_int)
                    KR_curve_comp[1].append(K_int)
        KR_curve_comp = np.array(KR_curve_comp)

        # Porcentaje de tolerancia
        perc_tol = 0.03
        tang_mat = []
        for k in range(0, len(df_all)):
            a_0 = 0
            tang_point = ""
            Tangent_Point = "No"
            Limit_Reach = "No"
            if len(df_all[k] > 0):
                # Buscar si hay tangencia entre las curvas
                while Tangent_Point == "No" and Limit_Reach == "No":
                    for i in range(1, KR_curve_comp.shape[1]):
                        K_curve = KR_curve_comp[1, i]
                        a_curve = KR_curve_comp[0, i] + a_0
                        K_actual = 0
                        for j in range(0, len(df_all[k]["Crack a"]) - 1):
                            if a_curve > df_all[k]["Crack a"][j] and a_curve < df_all[k]["Crack a"][j + 1]:
                                K_actual = df_all[k]["Klim ab"][j] + (df_all[k]["Klim ab"][j + 1] - df_all[k]["Klim ab"][j])/(df_all[k]["Crack a"][j + 1] - df_all[k]["Crack a"][j]) *(a_curve - df_all[k]["Crack a"][j])
                                break

                        # Si la K actual está en este rango de valores significa que hay tangencia (con tolerancia)
                        if K_actual > K_curve * (1 - perc_tol) and K_actual < K_curve * (1 + perc_tol):
                            if Tangent_Point == "Yes": # THIS MEANS THAT TWO TANGENT POINTS HAS BEEN FOUNDED WHICH CAN BE AN ERROR DUE TO THE TOLERANCE
                                Tangent_Point = "No"
                                break
                            else:
                                tang_point = a_0
                                Tangent_Point = "Yes"

                        # Si se rompe la tangencia es que la curva lo que hace es cortar
                        elif K_curve > K_actual and K_actual != 0:
                            Tangent_Point = "No"
                            break

                    # Si no se ha encontrado un punto de tangencia se va moviendo la curva R hacia la derecha
                    a_0 = a_0 + 1

                    if a_0 > df_all[0]["Crack a"][len(df_all[0]["Crack a"]) - 1]:
                        # Límite que se alcanza cuando a_0 es mayor que el máximo longitud de grieta obtenido
                        Limit_Reach = "Yes"
                        self.book["INPUTS"].cell(6, 25).value = "R-Curve tangent point found on final point"
                if Tangent_Point == "No":
                    # Si no ha habido punto de tangencia se hace la tangencia al final de la curva de los datos
                    tang_point = df_all[0]["Crack a"][len(df_all[0]["Crack a"]) - 1]
                tang_mat.append(tang_point)
            else:
                tang_mat.append("")
        return tang_mat


    def Residual_Strength_calc(self,df_all):
        # FUNCIÓN QUE CALCULA LA LONGITUD CRÍTICA DE GRIETA SEGÚN EL CRITERIO DE RESISTENCIA RESIDUAL
        crit_lengths = []
        for k in range(0, len(df_all)):
            if len(df_all[k] > 0):
                crit_len = "No"
                res_strength = []

                for i in range(0, len(df_all[0])):
                    # Se obtienen los valores de resistencia residual
                    if df_all[k]["Klim ab"][i] < df_all[k]["Klim cd"][i] and df_all[k]["Klim ab"][i] != 0:
                        res_strength.append(df_all[k]["Klim ab"][i])
                    elif df_all[k]["Klim cd"][i] < df_all[k]["Klim ab"][i] and df_all[k]["Klim cd"][i] != 0:
                        res_strength.append(df_all[k]["Klim cd"][i])
                    elif df_all[k]["Klim cd"][i] == 0:
                        res_strength.append(df_all[k]["Klim ab"][i])
                    elif df_all[k]["Klim ab"][i] == 0:
                        res_strength.append(df_all[k]["Klim cd"][i])

                    # Se comparan los valores de resistencia residual con la tensión límite para ver si se ha superado
                    if res_strength[i] < self.limit_stress:
                        crit_lengths.append(df_all[k]["Crack a"][i])
                        crit_len = "Yes"
                        break

                if crit_len == "No":
                    crit_lengths.append("")
            else:
                crit_lengths.append("")
        return crit_lengths


    def Net_Sec_Yield_calc(self,df_all, A_total, Fty):
        # FUNCIÓN QUE CALCULA LA LONGITUD CRÍTICA DE GRIETA SEGÚN EL CRITERIO DE SECCIÓN NETA
        crit_lengths = []

        Thickness = A_total / df_all[0]["Crack a"][len(df_all[0]["Crack a"]) - 1]
        for k in range(0, len(df_all)):
            if len(df_all[k] > 0):
                sig_residual = []
                sig_lim = []
                Crit_found = "No"
                for i in range(0, len(df_all[k])):
                    A_remaining = A_total - df_all[k]["a total"][i] * Thickness
                    sig_residual.append(Fty * A_remaining / A_total)

                    # Comparación de la resistencia de la sección neta con la tensión límite
                    if sig_residual[i] < self.limit_stress:
                        a_critical = df_all[k]["Crack a"][i - 1] + (df_all[k]["Crack a"][i] - df_all[k]["Crack a"][i - 1])/(sig_residual[i] - sig_residual[i - 1]) * (self.limit_stress - sig_residual[i - 1])
                        crit_lengths.append(a_critical)
                        Crit_found = "Yes"
                        break
                if Crit_found == "No":
                    crit_lengths.append("")
            else:
                crit_lengths.append("")
        return crit_lengths


    def Fast_Growth_Crack_calc(self, Method, df_all):
        # FUNCIÓN QUE CALCULA LA LONGITUD CRÍTICA DE GRIETA SEGÚN EL CRITERIO DE CRECIMIENTO RÁPIDO DE GRIETA

        crit_lengths = []
        for k in range(0, len(df_all)):
            if len(df_all[k] > 0):
                crack_growth = []
                Crit_found = "No"
                for i in range(0, len(df_all[k]) - 1):
                    # Ver qué método ha sido seleccionado y hacer los cálculos en base a ello
                    if Method == "1 mm/flight":
                        if (df_all[k]["Flights"][i + 1] - df_all[k]["Flights"][i]) == 0:
                            crack_growth.append(0)
                        else:
                            crack_growth.append((df_all[k]["Crack a"][i + 1] - df_all[k]["Crack a"][i])/(df_all[k]["Flights"][i + 1] - df_all[k]["Flights"][i]))
                    elif Method == "1 mm/cycle":
                        if (df_all[k]["Flights"][i + 1] - df_all[k]["Flights"][i]) == 0:
                            crack_growth.append(0)
                        else:
                            crack_growth.append((df_all[k]["Crack a"][i + 1] - df_all[k]["Crack a"][i]) / (df_all[k]["Cycles"][i + 1] - df_all[k]["Cycles"][i]))

                    if crack_growth[i] > 1:
                        crit_lengths.append(df_all[k]["Crack a"][i])
                        Crit_found = "Yes"
                        break
                if Crit_found == "No":
                    crit_lengths.append("")
            else:
                crit_lengths.append("")
        return crit_lengths


    def Write_Critical_Lengths(self, row, crit_lengths, Crit_crack_length_cons):
        # FUNCION PARA ESCRIBIR LOS VALORES DE LONGITUD CRÍTICA EN EL EXCEL
        col = 8
        for i in range(0, len(crit_lengths)):
            self.book["INPUTS"].cell(row, col).value = crit_lengths[i]
            col = col + 1

        # Obtener las máximas y mínimas longitudes críticas de grieta
        Min_length = crit_lengths[0]
        Max_length = crit_lengths[0]
        for i in range(1, len(crit_lengths)):
            if crit_lengths[i] != "":
                if crit_lengths[i] > Max_length:
                    Max_length = crit_lengths[i]
                elif crit_lengths[i] < Min_length:
                    Min_length = crit_lengths[i]

        # Escribir las longitudes críticas en el Excel
        if Crit_crack_length_cons == "Real":
            self.book["INPUTS"].cell(row, 13).value = Max_length
        elif Crit_crack_length_cons == "Conservative":
            self.book["INPUTS"].cell(row, 13).value = Min_length

        return Max_length, Min_length


    def Write_Critical_Lengths_txt(self, output_folder, filename, All_Crit_Lengths, Global_max_lengths, Global_min_lengths, Crit_crack_length_cons):
        # FUNCIÓN PARA ESCRIBIR LOS DATOS DE LONGITUDES CRÍTICAS EN UN ARCHIVO DE TEXTO

        # Creacíon del archivo y escritura de los datos iniciales
        output_file = output_folder + "/" + filename
        if os.path.isfile(output_file + '.txt'):  # Creación del archivo txt de salida.
            os.remove(output_file + '.txt')
        file = open(output_file + '.txt', "x")
        file.writelines("##################################\n") # Encabezado del archivo de entrada a ISAMI.
        file.writelines("#      GRIETA VERSION x.x        #\n")
        file.writelines("#  GRIETA_READER VERSION: v1.0   #\n")
        file.writelines("##################################\n")
        file.writelines("\n")
        file.writelines("\n")
        file.writelines("CRITICAL LENGTHS RESULTS:\n")
        file.writelines("\n")

        # Creación de una tabla mediante librería PrettyTable para poder escribirla en el txt
        Lengths_Table = PrettyTable()
        field_names = ["", "SR", "MR", "LR", "ULR", "MIX", "Lcrit (mm)", "Lcrit global (mm)"]
        headers = ["Fracture Mechanics", "Net Section Yield", "Fast Crack Growth"]
        Lengths_Table.field_names = field_names

        # Se pone en vectores las máximas y mínimas para poder meterlas en la tabla
        for i in range(0, len(Global_max_lengths)):
            if Global_max_lengths[i] != "":
                Max_length = Global_max_lengths[i]
                Min_length = Global_min_lengths[i]
                break
        for i in range(1, len(Global_max_lengths)):
            if Global_max_lengths[i] != "":
                if Global_max_lengths[i] > Max_length:
                    Max_length = Global_max_lengths[i]
            if Global_min_lengths[i] != "":
                if Global_min_lengths[i] < Min_length:
                    Min_length = Global_min_lengths[i]

        row_lines = []
        for i in range(0, len(All_Crit_Lengths)):
            row_lines.append([])
            row_lines[i].append(headers[i])
            for j in range(0, len(All_Crit_Lengths[i])):
                row_lines[i].append(All_Crit_Lengths[i][j])

            if Crit_crack_length_cons == "Real":
                row_lines[i].append(Global_max_lengths[i])
                if i == 1:
                    row_lines[i].append(Max_length)
                else:
                    row_lines[i].append("")
            elif Crit_crack_length_cons == "Conservative":
                row_lines[i].append(Global_min_lengths[i])
                if i == 1:
                    row_lines[i].append(Min_length)
                else:
                    row_lines[i].append("")

        # Se añaden las líneas y se escribe en el txt
        Lengths_Table.add_rows(row_lines)
        file.write(str(Lengths_Table))
        file.writelines("\n")
        file.close()


