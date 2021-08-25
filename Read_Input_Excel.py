import os
from typing import List, Any

import numpy as np
from path import Path
import pandas as pd
from matplotlib import pyplot as plt
from openpyxl import load_workbook
from wrappers import read_table_horiz

class Read_Input_Excel():
    """
    Clase creada para leer el input del archivo Excel
    """
    def __init__(self, Excel_file, folder_data):
        self.folder_data = folder_data
        self.Excel_file = Excel_file
        self.book = load_workbook(self.Excel_file, data_only = True)
        self.files_missions = []
        for row in range(7,12):
            self.files_missions.append(self.book["INPUTS"].cell(row,2).value)
        self.Crack_Length_Calc_Method = self.book["INPUTS"].cell(14,2).value
        self.hole_diameter = self.book["INPUTS"].cell(17,2).value
        self.limit_stress = self.book["INPUTS"].cell(21,2).value
        self.load_case = self.book["INPUTS"].cell(22,2).value


    def Read_files(self):
        list_files = os.listdir(self.folder_data)
        SR_files = []
        MR_files = []
        LR_files = []
        ULR_files = []
        MIX_files = []
        for file in list_files:
            for mission_file in self.files_missions:
                if mission_file != None:
                    if file[:len(mission_file)] == mission_file and file[-5:] == "num.1":
                        if mission_file[-2:] == "sr":
                            SR_files.append(file)
                        elif mission_file[-2:] == "mr":
                            MR_files.append(file)
                        elif mission_file[-2:] == "lr":
                            LR_files.append(file)
                        elif mission_file[-3:] == "ulr":
                            ULR_files.append(file)
                        elif mission_file[-3:] == "mix":
                            MIX_files.append(file)
        All_missions = []
        All_missions.append(SR_files)
        All_missions.append(MR_files)
        All_missions.append(LR_files)
        All_missions.append(ULR_files)
        All_missions.append(MIX_files)

        content_mat = []
        for i in range(0, len(All_missions)):
            content_mat.append([])
            input_steps = 0
            for j in range(0, len(All_missions[i])):
                filename = self.folder_data + "/" + All_missions[i][j]
                with open(filename) as f:
                    content = f.readlines()
                    for line in content:
                        new_list_2 = []
                        new_list = line.split()
                        for item in new_list:
                            new_list_2.append(float(item))

                        All_zeros = "Yes"
                        for t in range(0, len(new_list_2)):
                            if new_list_2[t] != 0:
                                All_zeros = "No"

                        if All_zeros == "No":
                            new_list_2.append(input_steps)
                            content_mat[i].append(new_list_2)
                        else:
                            input_steps = input_steps + 1
        content = content_mat
        # Reroder content list
        content_order = []
        for k in range(0, len(content)):
            content_order.append([])
            if len(content[k]) > 0:
                for j in range(0, len(content[k][0])):
                    content_order[k].append([])
                    for i in range(0, len(content[k])):
                        content_order[k][j].append(content[k][i][j])

        name_list = ["Crack a", "Crack b", "Crack c", "Crack d", "Cycles", "Flights", "Beta a", "Beta b", "Beta c",
                     "Beta d", "Klim ab", "Klim cd", "Input Steps"]
        df_all = []
        for k in range(0, len(content_order)):
            df_all.append([])
            df_all[k] = pd.DataFrame()
            if len(content_order[k]) > 0:
                for i in range(0, len(name_list)):
                    df_all[k][name_list[i]] = content_order[k][i]

        return(df_all)


    def Obtain_Crack_Length(self, dataframe_all):
        for k in range(0, len(dataframe_all)):
            atotal_list = []
            for i in range(0, dataframe_all[k].shape[0]):
                if dataframe_all[k]["Crack b"][i] != 0 or dataframe_all[k]["Crack d"][i] != 0:
                    if self.Crack_Length_Calc_Method == "A":
                        atotal = dataframe_all[k]["Crack a"][i] + self.hole_diameter
                        atotal_list.append(atotal)
                    elif self.Crack_Length_Calc_Method == "c":
                        atotal = dataframe_all[k]["Crack c"][i] + self.hole_diameter
                        atotal_list.append(atotal)
                    elif self.Crack_Length_Calc_Method == "A+C":
                        atotal = dataframe_all[k]["Crack a"][i] + dataframe_all[k]["Crack c"][i] + self.hole_diameter
                        atotal_list.append(atotal)
                else:
                    if self.Crack_Length_Calc_Method == "A":
                        atotal = dataframe_all[k]["Crack a"][i]
                        atotal_list.append(atotal)
                    elif self.Crack_Length_Calc_Method == "c":
                        atotal = dataframe_all[k]["Crack c"][i]
                        atotal_list.append(atotal)
                    elif self.Crack_Length_Calc_Method == "A+C":
                        atotal = dataframe_all[k]["Crack a"][i] + dataframe_all[k]["Crack c"][i]
                        atotal_list.append(atotal)
            # dataframe.assign(atotal_list)
            dataframe_all[k]['a total'] = atotal_list
        return(dataframe_all)


    def Plot_Crack_Length(self, dataframe_all):
        missions_list = ["SR", "MR", "LR", "ULR", "MIX"]
        legend_mat = []
        for k in range(0, len(dataframe_all)):
            if len(dataframe_all[k]) > 0:
                plt.plot(dataframe_all[k]["Flights"], dataframe_all[k]['a total'])
                legend_mat.append(missions_list[k])
        plt.legend(legend_mat)
        plt.xlabel("Number of Flights")
        plt.ylabel("Crack length (mm)")
        plt.show()


    def Compute_Critical_Crack_Lengths(self, df_all):
        Fracture_Mechs_criterion = self.book["INPUTS"].cell(24,2).value
        Net_sect_yield_criterion = self.book["INPUTS"].cell(25, 2).value
        Fast_crack_growth_criterion = self.book["INPUTS"].cell(26, 2).value
        Crit_crack_length_cons = self.book["INPUTS"].cell(27, 2).value

        row_FM = 25
        row_NS = 26
        row_FC = 27

        Global_max_lengths = []
        Global_min_lengths = []
        # Obtain the critical lengths on each of the methods
        if Fracture_Mechs_criterion == "KR curve":
            # Read KR curve
            row = 35
            col = 1
            KR_curve_list = []
            cont = 0
            while self.book["INPUTS"].cell(row, col).value != None:
                KR_curve_list.append([])
                KR_curve_list[cont].append(self.book["INPUTS"].cell(row, col).value)
                KR_curve_list[cont].append(self.book["INPUTS"].cell(row, col + 1).value)
                row = row + 1
                cont = cont + 1

            KR_curve = np.array(KR_curve_list)
            Crit_Lengths_FM = self.KR_curve_calc(df_all, KR_curve)
            (Max_length_FM, Min_length_FM) = self.Write_Critical_Lengths(row_FM, Crit_Lengths_FM, Crit_crack_length_cons)
            if Min_length_FM != "":
                Global_min_lengths.append(Min_length_FM)

            if Max_length_FM != "":
                Global_max_lengths.append(Min_length_FM)

        elif Fracture_Mechs_criterion == "Residual strength":
            Crit_Lengths_FM = self.Residual_Strength_calc(df_all)
            (Max_length_FM, Min_length_FM) = self.Write_Critical_Lengths(row_FM, Crit_Lengths_FM,
                                                                         Crit_crack_length_cons)
            if Min_length_FM != "":
                Global_min_lengths.append(Min_length_FM)

            if Max_length_FM != "":
                Global_max_lengths.append(Min_length_FM)

        if Net_sect_yield_criterion == "Yes":
            A_total = self.book["INPUTS"].cell(29, 2).value
            Fty = self.book["INPUTS"].cell(30, 2).value
            Crit_Lengths_NSY = self.Net_Sec_Yield_calc(df_all, A_total, Fty)
            (Max_length_NSY, Min_length_NSY) = self.Write_Critical_Lengths(row_NS, Crit_Lengths_NSY, Crit_crack_length_cons)

            if Min_length_NSY != "":
                Global_min_lengths.append(Min_length_NSY)

            if Max_length_NSY != "":
                Global_max_lengths.append(Min_length_NSY)

        if Fast_crack_growth_criterion != "No":
            Crit_Lengths_CWC = self.Fast_Growth_Crack_calc(Fast_crack_growth_criterion, df_all)
            (Max_length_CWC, Min_length_CWC) = self.Write_Critical_Lengths(row_FC, Crit_Lengths_CWC, Crit_crack_length_cons)

            if Min_length_CWC != "":
                Global_min_lengths.append(Min_length_CWC)

            if Max_length_CWC != "":
                Global_max_lengths.append(Min_length_CWC)

        if Crit_crack_length_cons == "Real":
            self.book["INPUTS"].cell(25, 14).value = max(np.array(Global_max_lengths))
        elif Crit_crack_length_cons == "Conservative":
            self.book["INPUTS"].cell(25, 14).value = min(np.array(Global_min_lengths))

        self.book.save(self.Excel_file)


    def KR_curve_calc(self, df_all, KR_curve):
        des_inc = 5 # Minimum interval between two points of the R curve
        KR_curve_comp = []
        KR_curve_comp.append([])
        KR_curve_comp.append([])
        KR_curve_comp[0].append(KR_curve[0,0])
        KR_curve_comp[1].append(KR_curve[0,1])
        for i in range(1, len(KR_curve)):
            if KR_curve[i, 0] - KR_curve[i - 1, 0] <= des_inc:
                KR_curve_comp[0].append(KR_curve[i, 0])
                KR_curve_comp[1].append(KR_curve[i, 1])
            else:
                increments = round((KR_curve[i, 0] - KR_curve[i - 1, 0]) / des_inc)
                for j in range(0, increments):
                    a_int = KR_curve[i - 1, 0] + des_inc * (j + 1)
                    K_int = KR_curve[i - 1, 1] + (KR_curve[i, 1] - KR_curve[i - 1, 1])/(KR_curve[i, 0] - KR_curve[i - 1, 0]) * (a_int - KR_curve[i - 1, 0])
                    KR_curve_comp[0].append(a_int)
                    KR_curve_comp[1].append(K_int)
        KR_curve_comp = np.array(KR_curve_comp)

        perc_tol = 0.03
        tang_mat = []
        for k in range(0, len(df_all)):
            a_0 = 0
            tang_point = ""
            Tangent_Point = "No"
            Limit_Reach = "No"
            if len(df_all[k] > 0):
                while Tangent_Point == "No" and Limit_Reach == "No":
                    for i in range(1, KR_curve_comp.shape[1]):
                        K_curve = KR_curve_comp[1, i]
                        a_curve = KR_curve_comp[0, i] + a_0
                        K_actual = 0
                        for j in range(0, len(df_all[k]["Crack a"]) - 1):
                            if a_curve > df_all[k]["Crack a"][j] and a_curve < df_all[k]["Crack a"][j + 1]:
                                K_actual = df_all[k]["Klim ab"][j] + (df_all[k]["Klim ab"][j + 1] - df_all[k]["Klim ab"][j])/(df_all[k]["Crack a"][j + 1] - df_all[k]["Crack a"][j]) *(a_curve - df_all[k]["Crack a"][j])
                                break

                        if K_actual > K_curve * (1 - perc_tol) and K_actual < K_curve * (1 + perc_tol):
                            if Tangent_Point == "Yes": # THIS MEANS THAT TWO TANGENT POINTS HAS BEEN FOUNDED WHICH CAN BE AN ERROR DUE TO THE TOLERANCE
                                Tangent_Point = "No"
                                break
                            else:
                                tang_point = a_0
                                Tangent_Point = "Yes"

                        elif K_curve > K_actual and K_actual != 0:
                            Tangent_Point = "No"
                            break

                    a_0 = a_0 + 1
                    if a_0 > df_all[0]["Crack a"][len(df_all[0]["Crack a"]) - 1]:
                        Limit_Reach = "Yes"
                if Tangent_Point == "No":
                    tang_point = df_all[0]["Crack a"][len(df_all[0]["Crack a"]) - 1]
                tang_mat.append(tang_point)
            else:
                tang_mat.append("")
        return tang_mat


    def Residual_Strength_calc(self,df_all):
        crit_lengths = []
        for k in range(0, len(df_all)):
            if len(df_all[k] > 0):
                crit_len = "No"
                res_strength = []
                for i in range(0, len(df_all[0])):
                    if df_all[k]["Klim ab"][i] < df_all[k]["Klim cd"][i] and df_all[k]["Klim ab"][i] != 0:
                        res_strength.append(df_all[k]["Klim ab"][i])
                    elif df_all[k]["Klim cd"][i] < df_all[k]["Klim ab"][i] and df_all[k]["Klim cd"][i] != 0:
                        res_strength.append(df_all[k]["Klim cd"][i])
                    elif df_all[k]["Klim cd"][i] == 0:
                        res_strength.append(df_all[k]["Klim ab"][i])
                    elif df_all[k]["Klim ab"][i] == 0:
                        res_strength.append(df_all[k]["Klim cd"][i])

                    #if df_all[k]["Beta a"][i] < df_all[k]["Beta b"][i] and df_all[k]["Beta a"][i] != 0:
                    #    res_strength.append(
                    #        df_all[k]["Klim ab"][i] / np.sqrt(np.pi * df_all[k]["Crack a"][i]) / df_all[k]["Beta a"][i])
                    #elif df_all[k]["Beta b"][i] < df_all[k]["Beta a"][i] and df_all[k]["Beta b"][i] != 0:
                    #    res_strength.append(
                    #        df_all[k]["Klim ab"][i] / np.sqrt(np.pi * df_all[k]["Crack a"][i]) / df_all[k]["Beta b"][i])
                    #elif df_all[k]["Beta b"][i] == 0:
                    #    res_strength.append(
                    #        df_all[k]["Klim ab"][i] / np.sqrt(np.pi * df_all[k]["Crack a"][i]) / df_all[k]["Beta a"][i])
                    #elif df_all[k]["Beta a"][i] == 0:
                    #    res_strength.append(
                    #        df_all[k]["Klim ab"][i] / np.sqrt(np.pi * df_all[k]["Crack a"][i]) / df_all[k]["Beta b"][i])



                    if res_strength[i] < self.limit_stress:
                        crit_lengths.append(df_all[k]["Crack a"][i])
                        crit_len = "Yes"
                        break
                if crit_len == "No":
                    crit_lengths.append("")
            else:
                crit_lengths.append("")
        return crit_lengths


    def Net_Sec_Yield_calc(self,df_all, A_total, Fty):
        crit_lengths = []
        Thickness = A_total / df_all[0]["Crack a"][len(df_all[0]["Crack a"]) - 1]
        for k in range(0, len(df_all)):
            if len(df_all[k] > 0):
                sig_residual = []
                sig_lim = []
                Crit_found = "No"
                for i in range(0, len(df_all[k])):
                    A_remaining = A_total - df_all[k]["a total"][i] * Thickness
                    sig_residual.append(Fty * A_remaining / A_total)
                    if sig_residual[i] < self.limit_stress:
                        a_critical = df_all[k]["Crack a"][i - 1] + (df_all[k]["Crack a"][i] - df_all[k]["Crack a"][i - 1])/(sig_residual[i] - sig_residual[i - 1]) * (self.limit_stress - sig_residual[i - 1])
                        crit_lengths.append(a_critical)
                        Crit_found = "Yes"
                        break
                if Crit_found == "No":
                    crit_lengths.append("")
            else:
                crit_lengths.append("")
        return crit_lengths


    def Fast_Growth_Crack_calc(self, Method, df_all):
        crit_lengths = []
        for k in range(0, len(df_all)):
            if len(df_all[k] > 0):
                crack_growth = []
                Crit_found = "No"
                for i in range(0, len(df_all[k]) - 1):
                    if Method == "1 mm/flight":
                        if (df_all[k]["Flights"][i + 1] - df_all[k]["Flights"][i]) == 0:
                            crack_growth.append(0)
                        else:
                            crack_growth.append((df_all[k]["Crack a"][i + 1] - df_all[k]["Crack a"][i])/(df_all[k]["Flights"][i + 1] - df_all[k]["Flights"][i]))
                    elif Method == "1 mm/cycle":
                        if (df_all[k]["Flights"][i + 1] - df_all[k]["Flights"][i]) == 0:
                            crack_growth.append(0)
                        else:
                            crack_growth.append((df_all[k]["Crack a"][i + 1] - df_all[k]["Crack a"][i]) / (df_all[k]["Cycles"][i + 1] - df_all[k]["Cycles"][i]))

                    if crack_growth[i] > 1:
                        crit_lengths.append(df_all[k]["Crack a"][i])
                        Crit_found = "Yes"
                        break
                if Crit_found == "No":
                    crit_lengths.append("")
            else:
                crit_lengths.append("")
        return crit_lengths


    def Write_Critical_Lengths(self, row, crit_lengths, Crit_crack_length_cons):
        col = 8
        for i in range(0, len(crit_lengths)):
            self.book["INPUTS"].cell(row, col).value = crit_lengths[i]
            col = col + 1

        Min_length = crit_lengths[0]
        Max_length = crit_lengths[0]
        for i in range(1, len(crit_lengths)):
            if crit_lengths[i] != "":
                if crit_lengths[i] > Max_length:
                    Max_length = crit_lengths[i]
                elif crit_lengths[i] < Min_length:
                    Min_length = crit_lengths[i]

        if Crit_crack_length_cons == "Real":
            self.book["INPUTS"].cell(row, 13).value = Max_length
        elif Crit_crack_length_cons == "Conservative":
            self.book["INPUTS"].cell(row, 13).value = Min_length

        return Max_length, Min_length